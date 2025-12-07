#!/usr/bin/env python3
"""
Verify all answers in the IPv4 subnetting problems workbook
"""

import openpyxl
from openpyxl import load_workbook
import math

def ip_to_int(ip):
    """Convert IP address string to integer"""
    parts = [int(x) for x in ip.split('.')]
    return (parts[0] << 24) + (parts[1] << 16) + (parts[2] << 8) + parts[3]

def int_to_ip(num):
    """Convert integer to IP address string"""
    return f"{(num >> 24) & 0xFF}.{(num >> 16) & 0xFF}.{(num >> 8) & 0xFF}.{num & 0xFF}"

def cidr_to_mask(cidr):
    """Convert CIDR notation to subnet mask"""
    mask = (0xFFFFFFFF << (32 - cidr)) & 0xFFFFFFFF
    return int_to_ip(mask)

def mask_to_cidr(mask_str):
    """Convert subnet mask to CIDR notation"""
    mask_int = ip_to_int(mask_str)
    cidr = bin(mask_int).count('1')
    return cidr

def get_network_address(ip, cidr):
    """Get network address for an IP with given CIDR"""
    ip_int = ip_to_int(ip)
    mask = (0xFFFFFFFF << (32 - cidr)) & 0xFFFFFFFF
    return int_to_ip(ip_int & mask)

def verify_same_different_network(wb):
    """Verify Same/Different Network problems"""
    print("\n" + "="*80)
    print("VERIFYING: Same-Different Network Problems")
    print("="*80)

    ws = wb['Same-Different Network']
    errors = []

    for row_num in range(2, ws.max_row + 1):
        prob_num = ws[f'A{row_num}'].value
        if prob_num is None:
            break

        ip1 = ws[f'B{row_num}'].value
        ip2 = ws[f'C{row_num}'].value
        mask = ws[f'D{row_num}'].value
        cidr_str = ws[f'E{row_num}'].value
        answer = ws[f'G{row_num}'].value

        # Extract CIDR number
        cidr = int(cidr_str.replace('/', ''))

        # Calculate actual answer
        network1 = get_network_address(ip1, cidr)
        network2 = get_network_address(ip2, cidr)

        if network1 == network2:
            correct_answer = "Same network"
        else:
            correct_answer = "Different networks"

        # Verify
        if answer == correct_answer:
            print(f"✓ Problem {prob_num}: CORRECT - {ip1} and {ip2} with /{cidr} → {answer}")
        else:
            error_msg = f"✗ Problem {prob_num}: ERROR - {ip1} and {ip2} with /{cidr}"
            error_msg += f"\n  Expected: {correct_answer}, Got: {answer}"
            error_msg += f"\n  Network 1: {network1}, Network 2: {network2}"
            errors.append(error_msg)
            print(error_msg)

    return errors

def verify_common_subnet(wb):
    """Verify Common Subnet problems"""
    print("\n" + "="*80)
    print("VERIFYING: Common Subnet Problems")
    print("="*80)

    ws = wb['Common Subnet']
    errors = []

    for row_num in range(2, ws.max_row + 1):
        prob_num = ws[f'A{row_num}'].value
        if prob_num is None:
            break

        ip1 = ws[f'B{row_num}'].value
        subnet1_str = ws[f'C{row_num}'].value
        ip2 = ws[f'D{row_num}'].value
        subnet2_str = ws[f'E{row_num}'].value
        answer = ws[f'G{row_num}'].value

        # Parse CIDRs from subnet strings
        cidr1 = int(subnet1_str.split('/')[-1].strip(')'))
        cidr2 = int(subnet2_str.split('/')[-1].strip(')'))

        # Find common subnet by trying smaller CIDRs
        ip1_int = ip_to_int(ip1)
        ip2_int = ip_to_int(ip2)

        # Start from smallest CIDR and work down
        common_cidr = None
        for test_cidr in range(min(cidr1, cidr2) - 1, 0, -1):
            net1 = get_network_address(ip1, test_cidr)
            net2 = get_network_address(ip2, test_cidr)
            if net1 == net2:
                common_cidr = test_cidr
                common_network = net1
                break

        if common_cidr is None:
            errors.append(f"✗ Problem {prob_num}: ERROR - Could not find common subnet!")
            continue

        correct_answer = f"{common_network}/{common_cidr}"

        # Parse answer to compare (answer might include mask in parens)
        answer_network = answer.split(' ')[0]

        if answer_network == correct_answer:
            print(f"✓ Problem {prob_num}: CORRECT - Common subnet is {correct_answer}")
        else:
            error_msg = f"✗ Problem {prob_num}: ERROR"
            error_msg += f"\n  IP1: {ip1}/{cidr1}, IP2: {ip2}/{cidr2}"
            error_msg += f"\n  Expected: {correct_answer}, Got: {answer_network}"
            errors.append(error_msg)
            print(error_msg)

    return errors

def verify_subnet_division(wb):
    """Verify Subnet Division problems"""
    print("\n" + "="*80)
    print("VERIFYING: Subnet Division Problems")
    print("="*80)

    ws = wb['Subnet Division']
    errors = []
    row_num = 2

    prob_num = 1
    while row_num <= ws.max_row:
        if ws[f'A{row_num}'].value is None:
            row_num += 1
            continue

        if isinstance(ws[f'A{row_num}'].value, int):
            # This is a problem header row
            prob_num = ws[f'A{row_num}'].value
            base_network_str = ws[f'B{row_num}'].value
            num_subnets = ws[f'C{row_num}'].value
            new_cidr_str = ws[f'D{row_num}'].value
            new_mask = ws[f'E{row_num}'].value

            # Parse base network
            base_network, base_cidr_str = base_network_str.split('/')
            base_cidr = int(base_cidr_str)
            new_cidr = int(new_cidr_str.replace('/', ''))

            print(f"\nProblem {prob_num}: Dividing {base_network_str} into {num_subnets} subnets")

            # Verify subnet math
            subnet_bits = new_cidr - base_cidr
            calculated_subnets = 2 ** subnet_bits

            if calculated_subnets != num_subnets:
                error_msg = f"✗ Problem {prob_num}: Subnet count mismatch!"
                error_msg += f"\n  /{base_cidr} → /{new_cidr} should give {calculated_subnets} subnets, not {num_subnets}"
                errors.append(error_msg)
                print(error_msg)

            # Verify mask
            correct_mask = cidr_to_mask(new_cidr)
            if correct_mask != new_mask:
                error_msg = f"✗ Problem {prob_num}: Mask error! /{new_cidr} should be {correct_mask}, not {new_mask}"
                errors.append(error_msg)
                print(error_msg)

            # Skip to subnet details header
            row_num += 2

            # Verify each subnet
            subnet_size = 1 << (32 - new_cidr)
            base_net_int = ip_to_int(base_network)

            subnet_idx = 0
            while row_num <= ws.max_row and ws[f'B{row_num}'].value and isinstance(ws[f'B{row_num}'].value, int):
                subnet_num = ws[f'B{row_num}'].value
                network = ws[f'C{row_num}'].value
                first_host = ws[f'D{row_num}'].value
                last_host = ws[f'E{row_num}'].value
                num_hosts = ws[f'F{row_num}'].value

                # Calculate correct values
                subnet_net_int = base_net_int + (subnet_idx * subnet_size)
                correct_network = int_to_ip(subnet_net_int)
                correct_first = int_to_ip(subnet_net_int + 1)
                correct_last = int_to_ip(subnet_net_int + subnet_size - 2)
                correct_hosts = subnet_size - 2

                # Verify
                if network != correct_network:
                    error_msg = f"✗ Problem {prob_num}, Subnet {subnet_num}: Network should be {correct_network}, not {network}"
                    errors.append(error_msg)
                    print(error_msg)
                elif first_host != correct_first:
                    error_msg = f"✗ Problem {prob_num}, Subnet {subnet_num}: First host should be {correct_first}, not {first_host}"
                    errors.append(error_msg)
                    print(error_msg)
                elif last_host != correct_last:
                    error_msg = f"✗ Problem {prob_num}, Subnet {subnet_num}: Last host should be {correct_last}, not {last_host}"
                    errors.append(error_msg)
                    print(error_msg)
                elif num_hosts != correct_hosts:
                    error_msg = f"✗ Problem {prob_num}, Subnet {subnet_num}: Host count should be {correct_hosts}, not {num_hosts}"
                    errors.append(error_msg)
                    print(error_msg)
                else:
                    print(f"  ✓ Subnet {subnet_num}: {network} - {first_host} to {last_host} ({num_hosts} hosts)")

                subnet_idx += 1
                row_num += 1

        row_num += 1

    return errors

def verify_vlsm(wb):
    """Verify VLSM problems"""
    print("\n" + "="*80)
    print("VERIFYING: VLSM Problems")
    print("="*80)

    ws = wb['VLSM']
    errors = []
    row_num = 2

    while row_num <= ws.max_row:
        if ws[f'A{row_num}'].value is None:
            row_num += 1
            continue

        if isinstance(ws[f'A{row_num}'].value, int):
            prob_num = ws[f'A{row_num}'].value
            base_network_str = ws[f'B{row_num}'].value

            print(f"\nProblem {prob_num}: VLSM allocation from {base_network_str}")

            base_network, base_cidr_str = base_network_str.split('/')
            base_cidr = int(base_cidr_str)
            base_net_int = ip_to_int(base_network)

            # Skip to allocation header
            row_num += 2

            # Track allocations
            current_offset = 0
            dept_num = 0

            while row_num <= ws.max_row and ws[f'B{row_num}'].value and ws[f'B{row_num}'].value != '':
                if ws[f'B{row_num}'].value == 'Department':
                    row_num += 1
                    continue

                dept = ws[f'B{row_num}'].value
                hosts_needed = ws[f'C{row_num}'].value
                network = ws[f'D{row_num}'].value
                cidr_str = ws[f'E{row_num}'].value
                mask = ws[f'F{row_num}'].value
                first_host = ws[f'G{row_num}'].value
                last_host = ws[f'H{row_num}'].value
                usable_hosts = ws[f'I{row_num}'].value

                if dept is None or dept == '':
                    break

                dept_num += 1

                # Calculate minimum CIDR needed
                host_bits = math.ceil(math.log2(hosts_needed + 2))
                min_cidr = 32 - host_bits
                cidr = int(cidr_str.replace('/', ''))

                # Verify CIDR is sufficient
                subnet_size = 1 << (32 - cidr)
                available_hosts = subnet_size - 2

                if available_hosts < hosts_needed:
                    error_msg = f"✗ Problem {prob_num}, {dept}: Not enough hosts! /{cidr} only provides {available_hosts} hosts, needs {hosts_needed}"
                    errors.append(error_msg)
                    print(error_msg)

                # Calculate expected values
                subnet_net_int = base_net_int + current_offset
                correct_network = int_to_ip(subnet_net_int)
                correct_first = int_to_ip(subnet_net_int + 1)
                correct_last = int_to_ip(subnet_net_int + subnet_size - 2)
                correct_hosts = subnet_size - 2
                correct_mask = cidr_to_mask(cidr)

                # Verify
                errors_found = False
                if network != correct_network:
                    error_msg = f"✗ Problem {prob_num}, {dept}: Network should be {correct_network}, not {network}"
                    errors.append(error_msg)
                    print(error_msg)
                    errors_found = True

                if first_host != correct_first:
                    error_msg = f"✗ Problem {prob_num}, {dept}: First host should be {correct_first}, not {first_host}"
                    errors.append(error_msg)
                    print(error_msg)
                    errors_found = True

                if last_host != correct_last:
                    error_msg = f"✗ Problem {prob_num}, {dept}: Last host should be {correct_last}, not {last_host}"
                    errors.append(error_msg)
                    print(error_msg)
                    errors_found = True

                if usable_hosts != correct_hosts:
                    error_msg = f"✗ Problem {prob_num}, {dept}: Usable hosts should be {correct_hosts}, not {usable_hosts}"
                    errors.append(error_msg)
                    print(error_msg)
                    errors_found = True

                if mask != correct_mask:
                    error_msg = f"✗ Problem {prob_num}, {dept}: Mask should be {correct_mask}, not {mask}"
                    errors.append(error_msg)
                    print(error_msg)
                    errors_found = True

                if not errors_found:
                    print(f"  ✓ {dept}: {network}/{cidr} ({hosts_needed} needed, {usable_hosts} available)")

                current_offset += subnet_size
                row_num += 1

        row_num += 1

    return errors

def verify_supernetting(wb):
    """Verify Supernetting/Route Summarization problems"""
    print("\n" + "="*80)
    print("VERIFYING: Supernetting/Route Summarization Problems")
    print("="*80)

    ws = wb['Supernetting']
    errors = []
    row_num = 2

    while row_num <= ws.max_row:
        prob_num = ws[f'A{row_num}'].value
        if prob_num is None or not isinstance(prob_num, int):
            row_num += 1
            continue

        networks_str = ws[f'B{row_num}'].value
        summary_network = ws[f'D{row_num}'].value
        summary_mask = ws[f'E{row_num}'].value
        summary_cidr_str = ws[f'F{row_num}'].value

        summary_cidr = int(summary_cidr_str.replace('/', ''))

        print(f"\nProblem {prob_num}: Summary route {summary_network}/{summary_cidr}")

        # Skip to network list
        row_num += 2

        # Collect all networks
        networks = []
        while row_num <= ws.max_row and ws[f'B{row_num}'].value:
            value = ws[f'B{row_num}'].value
            if value and isinstance(value, str) and '/' in value:
                # Parse networks from comma-separated list
                net_list = [n.strip() for n in value.split(',') if '/' in n]
                networks.extend(net_list)
            row_num += 1

        if not networks:
            errors.append(f"✗ Problem {prob_num}: No networks found!")
            continue

        # Parse first and last network to verify summary
        first_net = networks[0].split('/')[0]
        first_cidr = int(networks[0].split('/')[1])
        last_net = networks[-1].split('/')[0]

        # Calculate expected summary
        first_net_int = ip_to_int(first_net)
        last_net_int = ip_to_int(last_net)

        # Find the common prefix
        # The summary should be the network address of the first network at the summary CIDR
        expected_summary = get_network_address(first_net, summary_cidr)

        # Verify all networks are within the summary
        summary_net_int = ip_to_int(summary_network)
        summary_size = 1 << (32 - summary_cidr)
        summary_last = summary_net_int + summary_size - 1

        all_within = True
        for net_str in networks:
            net_addr = net_str.split('/')[0]
            net_cidr = int(net_str.split('/')[1])
            net_int = ip_to_int(net_addr)

            if net_int < summary_net_int or net_int >= summary_net_int + summary_size:
                error_msg = f"✗ Problem {prob_num}: Network {net_str} is NOT within summary {summary_network}/{summary_cidr}"
                errors.append(error_msg)
                print(error_msg)
                all_within = False

        # Verify summary mask
        correct_mask = cidr_to_mask(summary_cidr)
        if summary_mask != correct_mask:
            error_msg = f"✗ Problem {prob_num}: Mask should be {correct_mask}, not {summary_mask}"
            errors.append(error_msg)
            print(error_msg)

        if all_within and summary_mask == correct_mask:
            print(f"  ✓ Summary {summary_network}/{summary_cidr} correctly contains all {len(networks)} networks")

        # Skip to next problem
        row_num += 1

    return errors

def main():
    """Main verification function"""
    print("\n" + "="*80)
    print("IPv4 SUBNETTING PROBLEMS - ANSWER VERIFICATION")
    print("="*80)

    wb = load_workbook("CSCI101_IPv4_Notes.xlsx")

    all_errors = []

    # Verify each problem type
    all_errors.extend(verify_same_different_network(wb))
    all_errors.extend(verify_common_subnet(wb))
    all_errors.extend(verify_subnet_division(wb))
    all_errors.extend(verify_vlsm(wb))
    all_errors.extend(verify_supernetting(wb))

    # Summary
    print("\n" + "="*80)
    print("VERIFICATION SUMMARY")
    print("="*80)

    if len(all_errors) == 0:
        print("✓ ALL PROBLEMS VERIFIED - No errors found!")
    else:
        print(f"✗ ERRORS FOUND: {len(all_errors)} total errors")
        print("\nError details:")
        for error in all_errors:
            print(f"\n{error}")

    print("\n" + "="*80)

if __name__ == "__main__":
    main()
