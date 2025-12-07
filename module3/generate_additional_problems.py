#!/usr/bin/env python3
"""
Generate additional IPv4 subnetting problems: VLSM and Supernetting
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import random
import math

def ip_to_int(ip):
    """Convert IP address string to integer"""
    parts = [int(x) for x in ip.split('.')]
    return (parts[0] << 24) + (parts[1] << 16) + (parts[2] << 8) + parts[3]

def int_to_ip(num):
    """Convert integer to IP address string"""
    return f"{(num >> 24) & 0xFF}.{(num >> 16) & 0xFF}.{(num >> 8) & 0xFF}.{num & 0xFF}"

def cidr_to_mask(cidr):
    """Convert CIDR notation to subnet mask"""
    mask = (0xFFFFFFFF << (32 - cidr)) & 0xFFFFFFFF
    return int_to_ip(mask)

def get_network_address(ip, cidr):
    """Get network address for an IP with given CIDR"""
    ip_int = ip_to_int(ip)
    mask = (0xFFFFFFFF << (32 - cidr)) & 0xFFFFFFFF
    return int_to_ip(ip_int & mask)

def calculate_cidr_for_hosts(num_hosts):
    """Calculate minimum CIDR needed for given number of hosts"""
    # Need num_hosts + 2 (network + broadcast)
    host_bits = math.ceil(math.log2(num_hosts + 2))
    return 32 - host_bits

def generate_vlsm_problems(num_problems=10):
    """Generate VLSM problems with multiple departments/segments"""
    problems = []

    department_templates = [
        ['Sales', 'Engineering', 'HR', 'IT'],
        ['Production', 'Development', 'Testing', 'Management'],
        ['Finance', 'Marketing', 'Operations', 'Support'],
        ['Research', 'Design', 'Quality Assurance', 'Administration'],
        ['Customer Service', 'Warehouse', 'Shipping', 'Reception'],
        ['Lab A', 'Lab B', 'Server Room', 'Guest Network'],
        ['Building A', 'Building B', 'Building C', 'Remote Workers'],
        ['Manufacturing', 'Accounting', 'Legal', 'Training'],
        ['West Wing', 'East Wing', 'North Wing', 'South Wing'],
        ['Main Office', 'Branch Office', 'Data Center', 'WiFi Network']
    ]

    for i in range(num_problems):
        # Choose base network
        base_cidr = random.choice([20, 21, 22, 23])
        base_class = random.choice(['10', '172.16', '192.168'])

        if base_class == '10':
            base_ip = f"10.{random.randint(0, 255)}.0.0"
        elif base_class == '172.16':
            base_ip = f"172.{random.randint(16, 31)}.0.0"
        else:
            base_ip = f"192.168.{random.randint(0, 255)}.0"

        base_network = get_network_address(base_ip, base_cidr)

        # Generate 3-4 departments with different host requirements
        dept_list = department_templates[i % len(department_templates)]
        num_depts = random.randint(3, min(4, len(dept_list)))
        departments = random.sample(dept_list, num_depts)

        # Generate host requirements (in descending order for efficient VLSM)
        host_counts = []
        for j in range(num_depts):
            if j == 0:
                # Largest department
                hosts = random.choice([250, 500, 1000, 120, 200])
            elif j == num_depts - 1:
                # Smallest department
                hosts = random.choice([5, 10, 14, 20, 25])
            else:
                # Medium departments
                hosts = random.choice([30, 50, 60, 100, 110])
            host_counts.append(hosts)

        # Sort departments by host count (descending) for VLSM
        dept_requirements = sorted(zip(departments, host_counts), key=lambda x: x[1], reverse=True)

        # Calculate subnets using VLSM
        base_net_int = ip_to_int(base_network)
        current_offset = 0
        allocations = []

        for dept_name, hosts_needed in dept_requirements:
            # Calculate required CIDR
            required_cidr = calculate_cidr_for_hosts(hosts_needed)

            # Calculate subnet details
            subnet_size = 1 << (32 - required_cidr)
            subnet_network = int_to_ip(base_net_int + current_offset)
            first_host = int_to_ip(base_net_int + current_offset + 1)
            last_host = int_to_ip(base_net_int + current_offset + subnet_size - 2)
            broadcast = int_to_ip(base_net_int + current_offset + subnet_size - 1)
            usable_hosts = subnet_size - 2
            subnet_mask = cidr_to_mask(required_cidr)

            allocations.append({
                'department': dept_name,
                'hosts_needed': hosts_needed,
                'network': subnet_network,
                'cidr': required_cidr,
                'mask': subnet_mask,
                'first_host': first_host,
                'last_host': last_host,
                'broadcast': broadcast,
                'usable_hosts': usable_hosts
            })

            current_offset += subnet_size

        problems.append({
            'base_network': f"{base_network}/{base_cidr}",
            'base_cidr': base_cidr,
            'departments': dept_requirements,
            'allocations': allocations
        })

    return problems

def generate_supernetting_problems(num_problems=10):
    """Generate supernetting/route summarization problems"""
    problems = []

    for i in range(num_problems):
        # Choose how many networks to summarize
        num_networks = random.choice([2, 4, 8, 16])

        # Choose base CIDR (the individual network size)
        base_cidr = random.choice([24, 25, 26, 27])

        # Calculate summary CIDR
        summary_bits = int(math.log2(num_networks))
        summary_cidr = base_cidr - summary_bits

        # Generate starting network
        if summary_cidr >= 16:
            # Class C style
            base_ip = f"192.168.{random.randint(0, 255 - num_networks)}.0"
        elif summary_cidr >= 12:
            # Class B style
            second_octet = random.randint(0, 255 - (num_networks if base_cidr >= 24 else num_networks * 256))
            base_ip = f"172.16.{second_octet}.0"
        else:
            # Class A style
            base_ip = f"10.{random.randint(0, 255)}.0.0"

        # Get the base network (aligned to summary boundary)
        base_network = get_network_address(base_ip, summary_cidr)
        base_net_int = ip_to_int(base_network)

        # Generate the individual networks
        network_size = 1 << (32 - base_cidr)
        networks = []
        for j in range(num_networks):
            network_addr = int_to_ip(base_net_int + (j * network_size))
            networks.append(f"{network_addr}/{base_cidr}")

        # Calculate summary route
        summary_network = base_network
        summary_mask = cidr_to_mask(summary_cidr)

        # Calculate summary route details
        summary_size = 1 << (32 - summary_cidr)
        summary_first = int_to_ip(base_net_int + 1)
        summary_last = int_to_ip(base_net_int + summary_size - 2)

        problems.append({
            'networks': networks,
            'num_networks': num_networks,
            'base_cidr': base_cidr,
            'summary_network': summary_network,
            'summary_cidr': summary_cidr,
            'summary_mask': summary_mask,
            'summary_route': f"{summary_network}/{summary_cidr}",
            'first_ip': summary_first,
            'last_ip': summary_last
        })

    return problems

def add_problems_to_excel(filename):
    """Add VLSM and Supernetting problems to the Excel file"""
    wb = load_workbook(filename)

    # Generate problems
    vlsm_problems = generate_vlsm_problems(10)
    supernetting_problems = generate_supernetting_problems(10)

    # Header styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True)

    # Sheet 4: VLSM Problems
    if 'VLSM' in wb.sheetnames:
        ws4 = wb['VLSM']
        wb.remove(ws4)
    ws4 = wb.create_sheet('VLSM')

    ws4.append(['Problem #', 'Base Network', 'Instructions'])
    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    current_row = 2
    for idx, prob in enumerate(vlsm_problems, 1):
        # Create requirements string
        requirements = []
        for dept, hosts in prob['departments']:
            requirements.append(f"{dept}: {hosts} hosts")
        requirements_str = "; ".join(requirements)

        instruction = f"Using VLSM, allocate subnets from {prob['base_network']} for the following departments: {requirements_str}"

        # Add problem header
        ws4.append([idx, prob['base_network'], instruction])
        ws4.row_dimensions[current_row].height = 40
        ws4[f'C{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')

        # Add solution header
        current_row += 1
        ws4.append(['', 'Department', 'Hosts Needed', 'Network Address', 'CIDR', 'Subnet Mask', 'First Host', 'Last Host', 'Usable Hosts'])
        for cell in ws4[current_row]:
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.alignment = Alignment(horizontal="center")

        # Add each allocation
        for alloc in prob['allocations']:
            current_row += 1
            ws4.append([
                '',
                alloc['department'],
                alloc['hosts_needed'],
                alloc['network'],
                f"/{alloc['cidr']}",
                alloc['mask'],
                alloc['first_host'],
                alloc['last_host'],
                alloc['usable_hosts']
            ])

        # Add spacing
        current_row += 2
        ws4.append([''])

    ws4.column_dimensions['A'].width = 12
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 70
    ws4.column_dimensions['D'].width = 18
    ws4.column_dimensions['E'].width = 8
    ws4.column_dimensions['F'].width = 18
    ws4.column_dimensions['G'].width = 18
    ws4.column_dimensions['H'].width = 18
    ws4.column_dimensions['I'].width = 15

    # Sheet 5: Supernetting/Route Summarization Problems
    if 'Supernetting' in wb.sheetnames:
        ws5 = wb['Supernetting']
        wb.remove(ws5)
    ws5 = wb.create_sheet('Supernetting')

    ws5.append(['Problem #', 'Networks to Summarize', 'Question', 'Summary Route', 'Summary Mask', 'CIDR'])
    for cell in ws5[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    current_row = 2
    for idx, prob in enumerate(supernetting_problems, 1):
        # Format networks list
        if len(prob['networks']) <= 4:
            networks_str = ", ".join(prob['networks'])
        else:
            # Show first few and last one
            networks_str = ", ".join(prob['networks'][:3]) + f", ... , {prob['networks'][-1]} ({len(prob['networks'])} networks total)"

        question = f"Find the summary route that encompasses all {prob['num_networks']} networks listed."

        ws5.append([
            idx,
            networks_str,
            question,
            prob['summary_network'],
            prob['summary_mask'],
            f"/{prob['summary_cidr']}"
        ])

        ws5.row_dimensions[current_row].height = 30
        ws5[f'B{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')

        # Add detailed network list as sub-rows
        current_row += 1
        ws5.append(['', 'Individual Networks:'])
        ws5[f'B{current_row}'].font = subheader_font

        # List all networks in chunks
        chunk_size = 4
        for i in range(0, len(prob['networks']), chunk_size):
            chunk = prob['networks'][i:i+chunk_size]
            current_row += 1
            ws5.append(['', ", ".join(chunk)])

        # Add spacing
        current_row += 2
        ws5.append([''])

    ws5.column_dimensions['A'].width = 12
    ws5.column_dimensions['B'].width = 50
    ws5.column_dimensions['C'].width = 45
    ws5.column_dimensions['D'].width = 20
    ws5.column_dimensions['E'].width = 20
    ws5.column_dimensions['F'].width = 10

    # Save the workbook
    wb.save(filename)
    print(f"Successfully added 20 additional problems to {filename}")
    print(f"  - 10 VLSM problems")
    print(f"  - 10 Supernetting/Route Summarization problems")
    print(f"\nTotal problems in workbook: 50")

if __name__ == "__main__":
    add_problems_to_excel("CSCI101_IPv4_Notes.xlsx")
