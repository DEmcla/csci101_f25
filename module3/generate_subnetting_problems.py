#!/usr/bin/env python3
"""
Generate 30 IPv4 subnetting problems for CSCI 101 class
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import random

def ip_to_int(ip):
    """Convert IP address string to integer"""
    parts = [int(x) for x in ip.split('.')]
    return (parts[0] << 24) + (parts[1] << 16) + (parts[2] << 8) + parts[3]

def int_to_ip(num):
    """Convert integer to IP address string"""
    return f"{(num >> 24) & 0xFF}.{(num >> 16) & 0xFF}.{(num >> 8) & 0xFF}.{num & 0xFF}"

def cidr_to_mask(cidr):
    """Convert CIDR notation to subnet mask"""
    mask = (0xFFFFFFFF << (32 - cidr)) & 0xFFFFFFFF
    return int_to_ip(mask)

def get_network_address(ip, cidr):
    """Get network address for an IP with given CIDR"""
    ip_int = ip_to_int(ip)
    mask = (0xFFFFFFFF << (32 - cidr)) & 0xFFFFFFFF
    return int_to_ip(ip_int & mask)

def generate_random_ip(network_class='any'):
    """Generate a random IP address"""
    if network_class == 'A':
        return f"{random.randint(1, 126)}.{random.randint(0, 255)}.{random.randint(0, 255)}.{random.randint(1, 254)}"
    elif network_class == 'B':
        return f"{random.randint(128, 191)}.{random.randint(0, 255)}.{random.randint(0, 255)}.{random.randint(1, 254)}"
    elif network_class == 'C':
        return f"{random.randint(192, 223)}.{random.randint(0, 255)}.{random.randint(0, 255)}.{random.randint(1, 254)}"
    else:
        return f"{random.randint(1, 223)}.{random.randint(0, 255)}.{random.randint(0, 255)}.{random.randint(1, 254)}"

def generate_same_network_problems(num_problems=10):
    """Generate problems asking if two IPs are on same network"""
    problems = []

    for i in range(num_problems):
        # Randomly decide if they should be on same network
        same_network = random.choice([True, False])
        cidr = random.choice([24, 25, 26, 27, 28, 16, 20, 22])

        ip1 = generate_random_ip()
        network = get_network_address(ip1, cidr)

        if same_network:
            # Generate second IP in same network
            net_int = ip_to_int(network)
            host_bits = 32 - cidr
            max_hosts = (1 << host_bits) - 2
            host_num = random.randint(1, max_hosts)
            ip2 = int_to_ip(net_int + host_num)
            answer = "Same network"
        else:
            # Generate second IP in different network
            ip2 = generate_random_ip()
            # Make sure it's actually different
            while get_network_address(ip2, cidr) == network:
                ip2 = generate_random_ip()
            answer = "Different networks"

        mask = cidr_to_mask(cidr)
        problems.append({
            'question': f"Are {ip1} and {ip2} on the same network with subnet mask {mask} (/{cidr})?",
            'ip1': ip1,
            'ip2': ip2,
            'mask': mask,
            'cidr': cidr,
            'answer': answer
        })

    return problems

def generate_common_subnet_problems(num_problems=10):
    """Generate problems asking to find common subnet for two IPs with different subnets"""
    problems = []

    for i in range(num_problems):
        # Generate a larger common network
        common_cidr = random.choice([16, 20, 22, 24])

        # Generate two smaller subnets within the common network
        subnet1_cidr = random.choice([cidr for cidr in [24, 25, 26, 27, 28] if cidr > common_cidr])
        subnet2_cidr = random.choice([cidr for cidr in [24, 25, 26, 27, 28] if cidr > common_cidr])

        # Generate base network
        base_ip = generate_random_ip()
        common_network = get_network_address(base_ip, common_cidr)

        # Generate two IPs in different subnets but same common network
        common_net_int = ip_to_int(common_network)
        common_host_bits = 32 - common_cidr
        max_offset = (1 << common_host_bits) - 2

        offset1 = random.randint(1, max_offset // 2)
        offset2 = random.randint(max_offset // 2 + 1, max_offset)

        ip1 = int_to_ip(common_net_int + offset1)
        ip2 = int_to_ip(common_net_int + offset2)

        mask1 = cidr_to_mask(subnet1_cidr)
        mask2 = cidr_to_mask(subnet2_cidr)
        common_mask = cidr_to_mask(common_cidr)

        problems.append({
            'question': f"Find the smallest common subnet that contains both {ip1}/{subnet1_cidr} and {ip2}/{subnet2_cidr}",
            'ip1': ip1,
            'subnet1': f"{mask1} (/{subnet1_cidr})",
            'ip2': ip2,
            'subnet2': f"{mask2} (/{subnet2_cidr})",
            'answer': f"{common_network}/{common_cidr} ({common_mask})"
        })

    return problems

def generate_subnet_division_problems(num_problems=10):
    """Generate problems asking to divide a network into subnets"""
    problems = []

    for i in range(num_problems):
        # Start with a base network
        base_cidr = random.choice([16, 20, 22, 24])
        base_ip = generate_random_ip()
        base_network = get_network_address(base_ip, base_cidr)

        # Calculate how many subnets we can create
        max_subnet_bits = 32 - base_cidr - 2  # Leave at least 2 bits for hosts
        if max_subnet_bits < 1:
            max_subnet_bits = 1

        # Choose number of subnet bits (which determines number of subnets)
        subnet_bits = random.randint(1, min(4, max_subnet_bits))
        num_subnets = 2 ** subnet_bits
        new_cidr = base_cidr + subnet_bits

        # Calculate subnet details
        base_net_int = ip_to_int(base_network)
        subnet_size = 1 << (32 - new_cidr)
        num_hosts = subnet_size - 2  # Subtract network and broadcast addresses

        subnets_info = []
        for j in range(min(num_subnets, 8)):  # Limit to first 8 subnets for brevity
            subnet_net_int = base_net_int + (j * subnet_size)
            network_addr = int_to_ip(subnet_net_int)
            first_host = int_to_ip(subnet_net_int + 1)
            last_host = int_to_ip(subnet_net_int + subnet_size - 2)
            broadcast = int_to_ip(subnet_net_int + subnet_size - 1)

            subnets_info.append({
                'network': network_addr,
                'first_host': first_host,
                'last_host': last_host,
                'broadcast': broadcast,
                'num_hosts': num_hosts
            })

        new_mask = cidr_to_mask(new_cidr)

        problems.append({
            'question': f"Divide the network {base_network}/{base_cidr} into {num_subnets} subnets. For each subnet, identify the network address, first host, last host, and number of hosts.",
            'base_network': f"{base_network}/{base_cidr}",
            'num_subnets': num_subnets,
            'new_cidr': new_cidr,
            'new_mask': new_mask,
            'subnets': subnets_info
        })

    return problems

def add_problems_to_excel(filename):
    """Add all problems to the Excel file"""
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

    # Create sheets for each problem type
    sheet_names = ['Same-Different Network', 'Common Subnet', 'Subnet Division']

    # Generate problems
    same_diff_problems = generate_same_network_problems(10)
    common_subnet_problems = generate_common_subnet_problems(10)
    division_problems = generate_subnet_division_problems(10)

    # Header styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Sheet 1: Same/Different Network Problems
    if 'Same-Different Network' in wb.sheetnames:
        ws1 = wb['Same-Different Network']
        ws1.delete_rows(1, ws1.max_row)
    else:
        ws1 = wb.create_sheet('Same-Different Network')

    ws1.append(['Problem #', 'IP Address 1', 'IP Address 2', 'Subnet Mask', 'CIDR', 'Question', 'Answer'])
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for idx, prob in enumerate(same_diff_problems, 1):
        ws1.append([
            idx,
            prob['ip1'],
            prob['ip2'],
            prob['mask'],
            f"/{prob['cidr']}",
            prob['question'],
            prob['answer']
        ])

    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 8
    ws1.column_dimensions['F'].width = 60
    ws1.column_dimensions['G'].width = 20

    # Sheet 2: Common Subnet Problems
    if 'Common Subnet' in wb.sheetnames:
        ws2 = wb['Common Subnet']
        ws2.delete_rows(1, ws2.max_row)
    else:
        ws2 = wb.create_sheet('Common Subnet')

    ws2.append(['Problem #', 'IP Address 1', 'Subnet 1', 'IP Address 2', 'Subnet 2', 'Question', 'Answer'])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for idx, prob in enumerate(common_subnet_problems, 1):
        ws2.append([
            idx,
            prob['ip1'],
            prob['subnet1'],
            prob['ip2'],
            prob['subnet2'],
            prob['question'],
            prob['answer']
        ])

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 20
    ws2.column_dimensions['F'].width = 60
    ws2.column_dimensions['G'].width = 35

    # Sheet 3: Subnet Division Problems
    if 'Subnet Division' in wb.sheetnames:
        ws3 = wb['Subnet Division']
        ws3.delete_rows(1, ws3.max_row)
    else:
        ws3 = wb.create_sheet('Subnet Division')

    ws3.append(['Problem #', 'Base Network', 'Subnets Requested', 'New CIDR', 'New Subnet Mask', 'Question'])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    current_row = 2
    for idx, prob in enumerate(division_problems, 1):
        # Add problem header
        ws3.append([
            idx,
            prob['base_network'],
            prob['num_subnets'],
            f"/{prob['new_cidr']}",
            prob['new_mask'],
            prob['question']
        ])

        # Add subnet details header
        current_row += 1
        ws3.append(['', 'Subnet #', 'Network Address', 'First Host', 'Last Host', 'Hosts per Subnet'])
        for cell in ws3[current_row]:
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.font = Font(bold=True)

        # Add each subnet's details
        for subnet_idx, subnet in enumerate(prob['subnets'], 1):
            current_row += 1
            ws3.append([
                '',
                subnet_idx,
                subnet['network'],
                subnet['first_host'],
                subnet['last_host'],
                subnet['num_hosts']
            ])

        # Add empty row for spacing
        current_row += 2
        ws3.append([''])

    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 18
    ws3.column_dimensions['E'].width = 20
    ws3.column_dimensions['F'].width = 60

    # Save the workbook
    wb.save(filename)
    print(f"Successfully added 30 subnetting problems to {filename}")
    print(f"  - 10 Same/Different Network problems")
    print(f"  - 10 Common Subnet problems")
    print(f"  - 10 Subnet Division problems")

if __name__ == "__main__":
    add_problems_to_excel("CSCI101_IPv4_Notes.xlsx")
